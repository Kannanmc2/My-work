import json
from openpyxl import load_workbook
from collections import OrderedDict
import collections
filepath = "test.xlsx"
d = collections.defaultdict(dict)
json_data = []
value = dict()
final_res = dict()
wb = load_workbook(filepath)
sheet = wb.active
mrow = sheet.max_row
mcolumn = sheet.max_column
for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
    json_data.append([x for x in [cell.value for cell in row] if x is not None])

for k, j in enumerate(range(0,len(json_data[2]), len(set(json_data[2])))):
    for i in json_data[3:]:
        value.setdefault(json_data[1][k], []).append(OrderedDict({stocks: prices for stocks, prices in zip([i for n, i in enumerate(json_data[2]) if i not in json_data[2][:n]], i[j:(j + len(set(json_data[2])))])}))
final_res.setdefault(json_data[0][0], []).append(value)
json_object = json.dumps(final_res)
print(json_object)
